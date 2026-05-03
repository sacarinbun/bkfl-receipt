"""
BKFL Receipt Processor — Streamlit Web App
อัปโหลด APSx export + bank statements → ดาวน์โหลด Excel คำนวณ

Usage:
    streamlit run app.py
"""

import io
import json
import re
import os
import datetime
import tempfile
from copy import copy
from pathlib import Path

import streamlit as st
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string as col_idx, range_boundaries
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    import pytesseract
    from pdf2image import convert_from_path
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

THAI_MONTHS = {
    'ม.ค.': '01', 'ก.พ.': '02', 'มี.ค.': '03', 'เม.ย.': '04',
    'พ.ค.': '05', 'มิ.ย.': '06', 'ก.ค.': '07', 'ส.ค.': '08',
    'ก.ย.': '09', 'ต.ค.': '10', 'พ.ย.': '11', 'ธ.ค.': '12',
}
EN_MONTHS = {
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
    'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
    'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12',
}
MDR_RATES = {'SCB_OOM': 0.0268, 'KBANK': 0.0257, 'BAY': 0.0214, 'SCB_DOEN': 0.0}

fill_scb_doen = PatternFill('solid', fgColor='D4AAFF')   # ม่วง
fill_scb_oom  = PatternFill('solid', fgColor='BDD7EE')   # น้ำเงินอ่อน
fill_kbank    = PatternFill('solid', fgColor='C6EFCE')   # เขียว
fill_bay      = PatternFill('solid', fgColor='FFEB9C')   # เหลือง
BANK_FILL     = {'SCB_DOEN': fill_scb_doen, 'SCB_OOM': fill_scb_oom,
                 'KBANK': fill_kbank, 'BAY': fill_bay}

BANK_ICON = {'SCB_DOEN': '🟪', 'SCB_OOM': '🟦', 'KBANK': '🟩', 'BAY': '🟨'}

SKIP_CODES = '{"A00","A01","A02","A03","A04","A05","A06","A07","A08","OGT"}'

PRODUCT_MAP = [
    ('SRH001','ดึงหน้า'),   ('SRH002','ดึงหน้า'),   ('SRH003','ดึงขมับ'),
    ('C2025004','ดึงหน้า'), ('C2025005','ดึงคอ'),   ('C2025006','ดึงคอ'),
    ('C2026007',''),
    ('INH001',''),
    ('SUB001','ตาสองชั้น'), ('SUB002','ตาสองชั้น'), ('SUB003','ตาสองชั้น'),
    ('SUB004','ตาสองชั้น'), ('SUB005','ตาสองชั้น'), ('SUB006','ตาสองชั้น'),
    ('SUB007','ตาสองชั้น'), ('SUB008','ตาสองชั้น'),
    ('SLB001','ตาล่าง'),    ('SLB002','ตาล่าง'),
    ('SRN001','จมูก'), ('SRN002','จมูก'), ('SRN003','จมูก'), ('SRN004','จมูก'),
    ('SRN005','จมูก'), ('SRN006','จมูก'), ('SRN007','จมูก'), ('SRN008','จมูก'),
    ('SCH001','เสริมคาง'),
    ('SFT001','ฉีดไขมัน'),  ('SFT002','ฉีดไขมัน'),  ('SFT003','ฉีดไขมัน'),
    ('SRF001','เสริมหน้าผาก'), ('SFH001','เสริมหน้าผาก'), ('C2026008','เสริมหน้าผาก'),
    ('SLR001','ปากกระจับ'), ('SLR002','ปากกระจับ'),
    ('SLS001','ปากไคลี่'),  ('SLS002','ปากไคลี่'),
    ('SBL001','ยกคิ้ว'),
    ('SFX001','foxy'),      ('SFX002','foxy'),
    ('SBA001','หน้าอก'),    ('TLS001','thread'),     ('SRE001','เคสแก้'),
    ('ABA001','botox'), ('ABA002','botox'), ('ABA003','botox'), ('ABA004','botox'),
    ('ABA005','botox'), ('ABA006','botox'), ('ABA007','botox'),
    ('ABE001','botox'), ('ABE002','botox'), ('ABE003','botox'), ('ABE004','botox'),
    ('ABE005','botox'), ('ABH006','botox'), ('ABH007','botox'),
    ('AFL001','filler'), ('GF001','filler'), ('FB001','filler'),
    ('SCT001','sculptra'), ('SCT003','sculptra'), ('SCT005','sculptra'),
    ('AUL001','ulthera'), ('AUL002','ulthera'),
    ('AMP001','morpheus'), ('AMP002','morpheus'),
    ('AHI001','HIFU'),    ('AHI002','HIFU'),
    ('AMS001','meso'),    ('AMS002','meso'),
    ('IPL001','skin'),    ('IPL003','skin'),
    ('C2025001','accutite'),
    ('ASC001','cell'), ('ASC002','cell'), ('ASC005','cell'), ('ACE001','cell'),
    ('AVT001','vitamin'),('AVT002','vitamin'),('AVT003','vitamin'),('AVT004','vitamin'),
    ('AVT005','vitamin'),('AVT006','vitamin'),('AVT007','vitamin'),('AVT008','vitamin'),
    ('AVT009','vitamin'),('AVT010','vitamin'),('AVT011','vitamin'),('AVT012','vitamin'),
    ('AVT013','vitamin'),('AVT014','vitamin'),('AVT015','vitamin'),('AVT016','vitamin'),
    ('AVT017','vitamin'),
    ('IV001','vitamin'), ('IV002','vitamin'),('IV003','vitamin'),('IV004','vitamin'),
    ('IV005','vitamin'), ('IV006','vitamin'),('IV007','vitamin'),('IV008','vitamin'),
    ('IVP001','vitamin'),('IVP002','vitamin'),('IVP003','vitamin'),
    ('C2025002','vitamin'),('TAI01','vitamin'),
    ('HPRP01','hair'),('HPRP05','hair'),('HPRP06','hair'),('HPRP10','hair'),('HPRP12','hair'),
    ('HSHF01','hair'),('HSHF03','hair'),('HSHF06','hair'),('HSHF09','hair'),('HSHF12','hair'),
    ('SIHR03','hair'),('SMSC01','hair'),
    ('STRIMAX01','hair'),('STRIMAX03','hair'),('STRIMAX06','hair'),
    ('STRIMAX09','hair'),('STRIMAX12','hair'),
    ('D2024001','hair'),('D2024002','hair'),('D2024003','hair'),('D2024004','hair'),
    ('D2024005','hair'),('D2024006','hair'),('D2024007','hair'),('D2024008','hair'),
    ('D2024009','hair'),('D2024010','hair'),
    ('D2025018','hair'),('D2025019','hair'),('D2025020','hair'),
    ('D2025015','skin'),('D2025016','skin'),('D2025017','skin'),
]

ref_data = [
    ('starting price', 15000, 'ดึงขมับ', 'hair'),
    ('step',           5000,  'ดึงบน',   'botox'),
    ('Sx1 weight',     2500,  'ดึงหน้าล่าง', 'ulthera'),
    ('Sx2 weight',     750,   'ดึงหน้า', 'vitamin'),
    ('start',          2,     'ตาสองชั้น','morpheus'),
    ('step',           0.5,   'ยกคิ้ว',  'skin'),
    ('max',            25000, 'จมูก',    'sculptra'),
    ('high-surg',      None,  'ดึงหน้า + ฉีดไขมัน', 'cell'),
    ('starting price', 100000,'ตาบน + ตาล่าง','accutite'),
    ('step',           50000, 'เสริมคาง','filler'),
    ('Sx1 weight',     2500,  'ฉีดไขมัน','HIFU'),
    ('Sx2 weight',     750,   'ยกปาก',   'meso'),
    ('start',          10,    'ยกมุมปาก', None),
    ('step',           5,     'ตาล่าง',   None),
    ('max',            200000,'เสริมหน้าผาก', None),
    ('non-surg',       None,   None,      None),
    ('df',             0.1,   None,      None),
    (None, None, 'thread', None),   (None, None, 'ปากกระจับ', None),
    (None, None, 'ปากไคลี่', None), (None, None, 'ดึงคอ', None),
    (None, None, 'foxy', None),     (None, None, 'หน้าอก', None),
    (None, None, 'เคสแก้', None),
]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def pf(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(',', '').strip())
    except: return None

def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font); dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment); dst.border = copy(src.border)
        dst.number_format = src.number_format

def detect_month_from_filename(name):
    """Extract YYYYMM from APSx Thai filename or English folder name."""
    for th, mm in THAI_MONTHS.items():
        m = re.search(rf'_(\d{{4}})\s*(?:ถึง|$)', name)
        if th in name and m:
            be_year = int(m.group(1))
            ce_year = be_year - 543
            return f"{ce_year}{mm}"
    n = name.lower()
    for mon, mm in EN_MONTHS.items():
        m = re.search(rf'{mon}(\d{{4}})', n)
        if m:
            return m.group(1) + mm
    m = re.search(r'(\d{6})', name)
    if m: return m.group(1)
    return datetime.datetime.now().strftime('%Y%m')

def detect_bank_type(fname):
    n = fname.lower()
    if 'scb' in n:
        if any(k in n for k in ['doen', 'เดิน', '_sa_', 'saving']): return 'SCB_DOEN'
        if any(k in n for k in ['oom', 'ออม', 'check', 'credit', 'edc', 'card']): return 'SCB_OOM'
        return 'SCB_DOEN'
    if any(k in n for k in ['kbank', 'กสิกร', 'kasikorn']): return 'KBANK'
    if any(k in n for k in ['krungsri', 'กรุงศรี', 'bay', 'ayudhya']): return 'BAY'
    return None

def detect_bank_from_channel(ch):
    if not ch: return None
    c = str(ch).lower()
    if 'เงินสด' in c or 'cash' in c: return 'CASH'
    if 'scb' in c or 'ไทยพาณิชย์' in c:
        if any(k in c for k in ['ออม', 'card', 'edc', 'บัตร']): return 'SCB_OOM'
        return 'SCB_DOEN'
    if any(k in c for k in ['kbank', 'กสิกร', 'kasikorn']): return 'KBANK'
    if any(k in c for k in ['krungsri', 'กรุงศรี', 'bay']): return 'BAY'
    if 'โอน' in c: return 'SCB_DOEN'
    return None

def parse_date_str(s):
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try: return datetime.datetime.strptime(str(s).strip(), fmt)
        except: pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# BANK PDF PARSERS
# ─────────────────────────────────────────────────────────────────────────────

def parse_scb(pdf_path, btype):
    if not HAS_PDF: return []
    credits, date_pat, amt_pat = [], re.compile(r'(\d{2}/\d{2}/\d{4})'), re.compile(r'([\d,]+\.\d{2})')
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                for line in (page.extract_text() or '').split('\n'):
                    dm = date_pat.match(line.strip())
                    if not dm: continue
                    dt = parse_date_str(dm.group(1))
                    if not dt: continue
                    amts = amt_pat.findall(line)
                    for a in amts[:-1]:
                        v = pf(a)
                        if v and v >= 100:
                            credits.append({'date': dt, 'amount': v, 'bank': btype, 'desc': line.strip()})
    except: pass
    return credits

def parse_kbank(pdf_path):
    if not HAS_PDF: return []
    credits, date_pat, amt_pat = [], re.compile(r'(\d{2}/\d{2}/\d{4})'), re.compile(r'([\d,]+\.\d{2})')
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.split('\n'):
                    dm = date_pat.search(line)
                    if not dm: continue
                    dt = parse_date_str(dm.group(1))
                    if not dt: continue
                    if any(k in line for k in ['รับเงินจากการขาย', 'edc', 'credit', 'settlement']):
                        amts = amt_pat.findall(line)
                        if amts:
                            v = pf(amts[0])
                            if v and v >= 100:
                                credits.append({'date': dt, 'amount': v, 'bank': 'KBANK', 'desc': line.strip()})
    except: pass
    return credits

def parse_bay_ocr(pdf_path):
    if not HAS_OCR: return []
    credits, date_pat, amt_pat = [], re.compile(r'(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})'), re.compile(r'([\d,]+\.\d{2})')
    try:
        images = convert_from_path(pdf_path, dpi=300)
        for img in images:
            text = pytesseract.image_to_string(img, lang='tha+eng')
            for line in text.split('\n'):
                if 'ACH' not in line: continue
                dm = date_pat.search(line)
                if not dm: continue
                dt = parse_date_str(dm.group(1))
                if not dt: continue
                amts = amt_pat.findall(line)
                if amts:
                    v = pf(amts[0])
                    if v and v >= 100:
                        credits.append({'date': dt, 'amount': v, 'bank': 'BAY', 'desc': line.strip()})
    except: pass
    return credits

PARSERS = {'SCB_DOEN': lambda p: parse_scb(p,'SCB_DOEN'),
           'SCB_OOM':  lambda p: parse_scb(p,'SCB_OOM'),
           'KBANK':    parse_kbank,
           'BAY':      parse_bay_ocr}

# ─────────────────────────────────────────────────────────────────────────────
# CORE PROCESSOR
# ─────────────────────────────────────────────────────────────────────────────

def analyze_source(ws):
    merged_a = {}
    for m in ws.merged_cells.ranges:
        if m.min_col <= 1 <= m.max_col:
            val = ws.cell(m.min_row, m.min_col).value
            for r in range(m.min_row, m.max_row + 1):
                merged_a[r] = val
    def get_a(r): return merged_a.get(r, ws.cell(r, 1).value)

    r1 = [ws.cell(1, c).value for c in range(1, 10)]
    non_null = [v for v in r1 if v is not None]
    has_title = len(non_null) <= 2 and non_null and isinstance(non_null[0], str) and len(non_null[0]) > 3
    hdr = 2 if has_title else 1
    dstart = hdr + 1
    dend, streak = dstart, 0
    for r in range(dstart, ws.max_row + 1):
        a = get_a(r)
        if a and str(a).startswith('RE'):
            dend, streak = r, 0
        else:
            streak += 1
            if streak > 15: break
    return has_title, hdr, dstart, dend

def build_merge_resolver(ws):
    mr = {}
    for rng in ws.merged_cells.ranges:
        mc, mr_, xc, xr = range_boundaries(str(rng))
        val = ws.cell(mr_, mc).value
        for r in range(mr_, xr + 1):
            for c in range(mc, xc + 1):
                mr[(r, c)] = val
    cont = set()
    for rng in ws.merged_cells.ranges:
        if rng.max_row > rng.min_row:
            for r in range(rng.min_row + 1, rng.max_row + 1):
                cont.add(r)
    return mr, cont

def auto_match(receipt_items, all_credits):
    bank_match, unmatched = {}, []
    for re_num, y_val, channel in receipt_items:
        btype = detect_bank_from_channel(channel)
        if btype in (None, 'CASH'): continue
        mdr = MDR_RATES.get(btype, 0)
        expected = float(y_val) * (1 - mdr)
        best_idx, best_diff = None, float('inf')
        for i, c in enumerate(all_credits):
            if c['bank'] != btype: continue
            diff = abs(c['amount'] - expected)
            rel  = diff / expected if expected else 1
            if rel <= 0.04 and diff < best_diff:
                best_diff, best_idx = diff, i
        if best_idx is not None:
            c = all_credits[best_idx]
            bank_match[(re_num, float(y_val))] = (c['amount'], c['date'], btype)
        else:
            unmatched.append({'re': re_num, 'amount': float(y_val), 'channel': channel,
                              'expected_net': expected, 'bank_type': btype})
    return bank_match, unmatched

def parse_override(override_dict):
    result = {}
    for key_str, val in override_dict.items():
        parts = key_str.rsplit('_', 1)
        if len(parts) != 2: continue
        try:
            re_num, y_val = parts[0], float(parts[1])
            amount = float(val[0])
            dt = parse_date_str(str(val[1])) or val[1]
            result[(re_num, y_val)] = (amount, dt, val[2])
        except: pass
    return result

def build_excel(src_bytes, bank_match, yyyymm, unmatched_deposits=None):
    """Build output workbook and return as bytes."""
    wb_src = load_workbook(io.BytesIO(src_bytes), data_only=True)
    ws_src = wb_src.active
    _, hdr, dstart, dend = analyze_source(ws_src)
    mr, cont_rows = build_merge_resolver(ws_src)
    MAP_LAST_ROW = 2 + len(PRODUCT_MAP)

    def gv(r, c): return mr.get((r, c), ws_src.cell(r, c).value)

    data_rows = []
    for r in range(dstart, dend + 1):
        row = [gv(r, c) for c in range(1, 30)]
        if r in cont_rows:
            for ci in range(19, 27): row[ci] = None
        data_rows.append((r, row))

    wb_out = Workbook()
    bold = Font(bold=True); hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    xmon_fill = PatternFill('solid', fgColor='FCE4D6'); umn_fill = PatternFill('solid', fgColor='E2EFDA')
    MONEY_FMT = '#,##0.00'
    DATE_FMT  = 'DD/MM/YYYY'

    # คำนวณ sheet (first)
    ws_c = wb_out.active; ws_c.title = 'คำนวณ'
    year = int(yyyymm[:4]); mon = int(yyyymm[4:])
    th_year = year + 543
    th_mon_names = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
                    'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    title = f'รายงานใบเสร็จ {th_mon_names[mon-1]} {th_year}'
    ws_c.cell(1, 1).value = title; ws_c.cell(1, 1).font = Font(bold=True, size=14)
    ws_c.merge_cells('A1:AW1')

    src_hdrs = [ws_src.cell(hdr, c).value for c in range(1, 30)]
    calc_hdrs = ['หัตถการ','การจ่าย','เงินเข้าบัญชี','วันที่เงินเข้า',
                 'ยอดสำหรับคิด DF/com','แพทย์1','DF1','วันที่จ่าย',
                 'แพทย์2','DF2','วันที่จ่าย','agency','%','ค่าคอม agency',
                 'วันที่จ่าย',''] + \
                ['low-surg', None, 'Surg-list', 'Non-surg-list']
    HEADERS = src_hdrs + calc_hdrs
    for ci, hv in enumerate(HEADERS, 1):
        dc = ws_c.cell(2, ci); dc.value = hv; dc.font = bold; dc.fill = hdr_fill

    ref_headers = {
        'AY':'รหัสสินค้า','AZ':'หัตถการ (keyword)',
        'BB':'ตารางมัดจำข้ามเดือน (กรอก manual)','BC':'HN',
        'BD':'ชื่อ-นามสกุล','BE':'หัตถการ','BF':'ยอดมัดจำค้าง DF','BG':'เดือนที่รับมัดจำ',
        'BI':'วันที่','BJ':'บัญชี','BK':'ยอด','BL':'หมายเหตุ',
    }
    _xmon_cols = {'BB','BC','BD','BE','BF','BG'}
    _umn_cols  = {'BI','BJ','BK','BL'}
    for ltr, lbl in ref_headers.items():
        ci = col_idx(ltr)
        f = xmon_fill if ltr in _xmon_cols else (umn_fill if ltr in _umn_cols else hdr_fill)
        ws_c.cell(2, ci).value = lbl; ws_c.cell(2, ci).fill = f; ws_c.cell(2, ci).font = bold
    ws_c.cell(1, 51).value = '▶ Mapping: รหัสสินค้า→หัตถการ'; ws_c.cell(1, 51).font = bold
    ws_c.cell(1, 54).value = '▶ มัดจำข้ามเดือน (กรอก manual)'; ws_c.cell(1, 54).font = bold
    ws_c.cell(1, 61).value = '▶ โอนธนาคารที่ยังไม่จับคู่';      ws_c.cell(1, 61).font = bold

    for i, row in enumerate(ref_data):
        ri = i + 3
        for j, v in enumerate(row): ws_c.cell(ri, 46 + j).value = v
    for i, (code, name) in enumerate(PRODUCT_MAP):
        ws_c.cell(i + 3, 51).value = code; ws_c.cell(i + 3, 52).value = name
    if unmatched_deposits:
        for i, (dt, bank, amt, note) in enumerate(unmatched_deposits):
            ri = i + 3
            ws_c.cell(ri, 61).value = dt; ws_c.cell(ri, 62).value = bank
            ws_c.cell(ri, 63).value = amt; ws_c.cell(ri, 64).value = note

    matched = 0; unmatched_cells = 0
    for idx, (r_src, row) in enumerate(data_rows):
        r = idx + 3
        for ci, val in enumerate(row, 1): ws_c.cell(r, ci).value = val

        # Col C: date only (no time)
        if isinstance(ws_c.cell(r, 3).value, datetime.datetime):
            ws_c.cell(r, 3).number_format = DATE_FMT
        # Col X: ยอดชำระ — money format
        ws_c.cell(r, 24).number_format = MONEY_FMT

        re_num  = row[0]         # col A
        y_val   = pf(row[23])    # col X
        pay_ch  = row[21]        # col V
        is_cont = (r_src in cont_rows)

        ws_c.cell(r, 30).value = (f'=IF(OR(L{r}={SKIP_CODES}),"-",'
                                   f'IFERROR(VLOOKUP(L{r},$AY$3:$AZ${MAP_LAST_ROW},2,FALSE),""))')
        ws_c.cell(r, 31).value = ''

        if is_cont:
            pass
        elif pay_ch and 'เงินสด' in str(pay_ch):
            ws_c.cell(r, 32).value = 0; ws_c.cell(r, 33).value = '-'
        elif y_val in (None, 0, 0.0):
            ws_c.cell(r, 32).value = 0; ws_c.cell(r, 33).value = '-'
        else:
            key = (re_num, float(y_val))
            bm  = bank_match.get(key)
            if bm:
                af, ag, btype = bm
                ws_c.cell(r, 32).value = af
                ws_c.cell(r, 32).fill  = BANK_FILL.get(btype, PatternFill())
                if isinstance(ag, datetime.datetime):
                    ws_c.cell(r, 33).value = ag; ws_c.cell(r, 33).number_format = DATE_FMT
                else:
                    ws_c.cell(r, 33).value = ag
                matched += 1
            else:
                ws_c.cell(r, 32).value = ''; ws_c.cell(r, 33).value = ''
                unmatched_cells += 1
        ws_c.cell(r, 32).number_format = MONEY_FMT

        ws_c.cell(r, 34).value = f'=IF(AE{r}="มัดจำ",0,IF(ISNUMBER(X{r}),X{r},0))'
        ws_c.cell(r, 34).number_format = MONEY_FMT
        ws_c.cell(r, 35).value = ''
        ws_c.cell(r, 36).value = (
            f'=IF(AH{r}=0,"-",IF(COUNTIF($AV$3:$AV$1045,$AD{r})>0,'
            f'IF($AH{r}<$AU$11,'
            f'IF(((QUOTIENT(($AH{r}-$AU$3),$AU$4)*$AU$8)+$AU$7)*$AU$5<=$AU$9,'
            f'((QUOTIENT(($AH{r}-$AU$3),$AU$4)*$AU$8)+$AU$7)*$AU$5,$AU$9),'
            f'IF(((QUOTIENT(($AH{r}-$AU$11),$AU$12)*$AU$16)+$AU$15)*$AU$13<=$AU$17,'
            f'((QUOTIENT(($AH{r}-$AU$11),$AU$12)*$AU$16)+$AU$15)*$AU$13,$AU$17)),'
            f'$AH{r}*$AU$19))')
        ws_c.cell(r, 36).number_format = MONEY_FMT
        ws_c.cell(r, 37).value = ''; ws_c.cell(r, 38).value = 'no'
        ws_c.cell(r, 39).value = (
            f'=IF($AH{r}=0,0,IF($AL{r}="no","no",'
            f'IF(COUNTIF($AV$3:$AV$1041,$AD{r})>0,'
            f'IF($AH{r}<$AU$11,'
            f'MIN(((QUOTIENT(($AH{r}-$AU$3),$AU$4)*$AU$8)+$AU$7)*$AU$6,$AU$9),'
            f'MIN(((QUOTIENT(($AH{r}-$AU$11),$AU$12)*$AU$16)+$AU$15)*$AU$14,$AU$17)),"no")))')
        ws_c.cell(r, 39).number_format = MONEY_FMT
        ws_c.cell(r, 40).value = 'no'
        ws_c.cell(r, 41).value = 'no'
        ws_c.cell(r, 42).value = f'=IF(AO{r}="no",0,IF(COUNTIF($AV$3:$AV$30,$AD{r})>0,10,5))'
        ws_c.cell(r, 43).value = f'=AH{r}*AP{r}%'
        ws_c.cell(r, 43).number_format = MONEY_FMT
        ws_c.cell(r, 44).value = 'no'
        ws_c.cell(r, 45).value = ''

    sum_row = dend + 3
    ws_c.cell(sum_row, 23).value = 'SUM ยอดชำระ →'; ws_c.cell(sum_row, 23).font = bold
    ws_c.cell(sum_row, 24).value = f'=SUM(X3:X{dend+1})'
    ws_c.cell(sum_row, 24).font = bold; ws_c.cell(sum_row, 24).number_format = MONEY_FMT
    ws_c.cell(sum_row, 31).value = 'SUM เงินเข้าบัญชี →'; ws_c.cell(sum_row, 31).font = bold
    ws_c.cell(sum_row, 32).value = f'=SUM(AF3:AF{dend+1})'
    ws_c.cell(sum_row, 32).font = bold; ws_c.cell(sum_row, 32).number_format = MONEY_FMT

    for cl, w in [('A',14),('B',14),('C',12),('G',12),('H',14),('L',12),('V',30),('X',12),
                  ('AD',15),('AE',13),('AF',14),('AG',13),('AH',16),('AI',13),('AJ',12),
                  ('AL',13),('AM',12),('AO',13),('AP',8),('AY',14),('AZ',18)]:
        ws_c.column_dimensions[cl].width = w
    for ci in range(54, 66): ws_c.column_dimensions[get_column_letter(ci)].width = 20
    ws_c.freeze_panes = 'A3'

    # Receipt Report sheet (second)
    ws_rr = wb_out.create_sheet('Receipt Report')
    wb_src2 = load_workbook(io.BytesIO(src_bytes), data_only=False)
    ws2 = wb_src2.active
    for r in range(1, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            sc = ws2.cell(r, c); dc = ws_rr.cell(r, c)
            dc.value = sc.value; copy_style(sc, dc)
    for m in ws2.merged_cells.ranges:
        ws_rr.merge_cells(start_row=m.min_row, start_column=m.min_col,
                          end_row=m.max_row, end_column=m.max_col)
    for cl, dim in ws2.column_dimensions.items():
        if dim.width: ws_rr.column_dimensions[cl].width = dim.width
    ws_rr.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue(), matched, unmatched_cells, len(data_rows), len(cont_rows)

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="BKFL Receipt Processor", page_icon="🏥", layout="wide")

st.title("🏥 BKFL Receipt Processor")
st.caption("อัปโหลดไฟล์จาก APSx + bank statements → ดาวน์โหลด Excel คำนวณ")

if not HAS_PDF:
    st.warning("⚠ pdfplumber ไม่ได้ติดตั้ง — bank PDF จะไม่ถูก parse อัตโนมัติ (ต้องใส่ override ทุกรายการ)")
if not HAS_OCR:
    st.info("ℹ pytesseract/pdf2image ไม่ได้ติดตั้ง — BAY PDF จะไม่ถูก parse ด้วย OCR")

# ── HELP ──────────────────────────────────────────────────────────────────────
with st.expander("📖 วิธีใช้งาน (คลิกเพื่อดู)"):
    st.markdown("""
### ขั้นตอนการใช้งาน

**ขั้นที่ 1 — เตรียมไฟล์**

| ไฟล์ | ได้มาจากไหน | ชื่อไฟล์ตัวอย่าง |
|------|------------|-----------------|
| APSx export (.xlsx) | www.clinic.app-apsx.com → ออกรายงาน → รายงานใบเสร็จ | รายงานใบเสร็จ_01_มี.ค._2569 ถึง 31_มี.ค._2569.xlsx |
| SCB เดิน (PDF) | K-Plus / SCB Easy → Statement | 03_SCBเดิน_HPPY_202603.PDF |
| SCB ออม / EDC (PDF) | SCB → Statement EDC/Card | 03_SCBออม_HPPY_202603.PDF |
| KBANK (PDF) | K-Plus → Statement | 03_Kbank_HPPY_202603.pdf |
| BAY กรุงศรี (PDF) | KMA → Statement | 03_krungsri_HPPY_202603.pdf |

> ไฟล์ bank PDF เป็น **optional** — ถ้าไม่มีก็ใส่ทุกรายการใน Override ด้านล่างแทนได้

---

**ขั้นที่ 2 — อัปโหลดและ Process**

1. อัปโหลดไฟล์ APSx ที่ช่อง **APSx Export** — ระบบจะตรวจเดือนจากชื่อไฟล์อัตโนมัติ
2. อัปโหลด bank PDF ทุกใบที่มีที่ช่อง **Bank Statements** (เลือกหลายไฟล์พร้อมกันได้)
3. กด **▶ Process**

---

**ขั้นที่ 3 — ดูผลลัพธ์**

หลัง Process ระบบจะแสดง metrics 4 ตัว:

| ตัวเลข | ความหมาย | ต้องทำอะไร |
|--------|---------|-----------|
| **Data rows** | จำนวนใบเสร็จทั้งหมด | — |
| **Sub-item rows zeroed** | แถวรายการย่อย (multi-product) ที่ถูก zero ไม่นับซ้ำ | — (อัตโนมัติ) |
| **AF/AG filled ✓** | ใบเสร็จที่ match กับ bank statement แล้ว | ยิ่งมากยิ่งดี |
| **Blank (unmatched) ⚠** | ใบเสร็จที่ยังไม่ match — AF/AG จะว่างใน Excel | ดูตารางด้านล่าง |

---

**ขั้นที่ 4 — แก้รายการที่ไม่ match (ถ้ามี)**

ถ้ามี **Blank > 0** ระบบจะแสดงตาราง พร้อม **Key สำหรับ Override** ในคอลัมน์สุดท้าย

คัดลอก Key และเปิด **Manual Override** box แล้ววาง JSON แบบนี้:
```
{
  "RE20261964_180000": [175176.0, "2026-04-03", "SCB_OOM"],
  "RE20261949_62319.95": [60452.84, "2026-03-02", "BAY"]
}
```
- ค่าแรก = **ยอดเงินจริงที่เข้าบัญชี** (ดูจาก bank statement)
- ค่าที่สอง = **วันที่เงินเข้า** (YYYY-MM-DD)
- ค่าที่สาม = **ประเภทบัญชี** → `SCB_DOEN` / `SCB_OOM` / `KBANK` / `BAY`

จากนั้นกด **▶ Process อีกครั้ง**

---

**กรณีพิเศษที่ปล่อยว่างได้ (ไม่ต้อง override)**

- **Cross-month settlement** — ใบเสร็จ 31 มี.ค. แต่เงินเข้าบัญชีเดือน เม.ย. → AF/AG ว่างได้ตามปกติ
- **เงินสด** — ระบบไม่กรอก AF/AG ให้ (ถูกต้องแล้ว)

---

**MDR rates อ้างอิง** (ยอดสุทธิที่เข้าบัญชีจริง)

| บัญชี | MDR | ยอดสุทธิ |
|------|-----|---------|
| SCB_DOEN (โอน) | 0% | เต็มยอด |
| SCB_OOM (card/EDC) | ~2.68% | × 0.9732 |
| KBANK | ~2.57% | × 0.9743 |
| BAY กรุงศรี | ~2.14% | × 0.9786 |
""")

# ── SECTION 1: UPLOAD ─────────────────────────────────────────────────────────
st.header("1. อัปโหลดไฟล์")
col1, col2 = st.columns([1, 1])

with col1:
    st.subheader("APSx Export (รายงานใบเสร็จ)")
    apsx_file = st.file_uploader("รายงานใบเสร็จ_*.xlsx จาก www.clinic.app-apsx.com",
                                  type=['xlsx'], key='apsx')
    if apsx_file:
        yyyymm = detect_month_from_filename(apsx_file.name)
        st.success(f"✓ ตรวจพบเดือน: **{yyyymm[:4]}/{yyyymm[4:]}**")

with col2:
    st.subheader("Bank Statements (PDFs)")
    bank_files = st.file_uploader("SCB เดิน / SCB ออม / KBANK / BAY กรุงศรี",
                                   type=['pdf', 'PDF'], accept_multiple_files=True, key='banks')
    if bank_files:
        for f in bank_files:
            btype = detect_bank_type(f.name)
            icon = BANK_ICON.get(btype, '⬜')
            st.write(f"{icon} **{btype or 'ไม่รู้จัก'}** — {f.name}")

# ── SECTION 2: MANUAL OVERRIDE ────────────────────────────────────────────────
with st.expander("🔧 Manual Override (bank_match_override.json)  —  กรอกถ้ามีรายการที่ match ไม่ได้"):
    st.caption(
        'Key format: `"RE20261949_62319.95"` → ใส่ยอดชำระจริง (รวมทศนิยม)\n\n'
        'Bank types: `SCB_DOEN` | `SCB_OOM` | `KBANK` | `BAY`'
    )
    sample = json.dumps({
        "RE20261949_62319.95": [60452.84, "2026-03-02", "BAY"],
        "RE20261939_1620": [1655.0, "2026-03-05", "SCB_DOEN"]
    }, ensure_ascii=False, indent=2)
    override_text = st.text_area("วาง JSON ที่นี่ (หรือเว้นว่างถ้าไม่มี)", height=200,
                                  placeholder=sample, key='override_input')

# ── SECTION 3: PROCESS ───────────────────────────────────────────────────────
st.header("2. ประมวลผล")
process_btn = st.button("▶ Process", type='primary', disabled=(apsx_file is None))

if process_btn and apsx_file:
    apsx_bytes = apsx_file.read()
    yyyymm = detect_month_from_filename(apsx_file.name)

    override_dict = {}
    if override_text and override_text.strip():
        try:
            override_dict = json.loads(override_text)
            st.success(f"✓ Loaded {len(override_dict)} manual overrides")
        except Exception as e:
            st.error(f"JSON parse error: {e}")

    all_credits = []
    with st.spinner("Parsing bank PDFs..."):
        for bf in (bank_files or []):
            btype = detect_bank_type(bf.name)
            if not btype:
                st.warning(f"⚠ ไม่รู้จักไฟล์: {bf.name} (ข้ามไป)")
                continue
            parser = PARSERS.get(btype)
            if not parser:
                continue
            with tempfile.NamedTemporaryFile(suffix=Path(bf.name).suffix, delete=False) as tmp:
                tmp.write(bf.read()); tmp_path = tmp.name
            try:
                credits = parser(tmp_path)
                all_credits.extend(credits)
                icon = BANK_ICON.get(btype, '⬜')
                st.write(f"  {icon} {btype}: {len(credits)} credit entries")
            finally:
                os.unlink(tmp_path)

    with st.spinner("Analyzing APSx source..."):
        wb_tmp = load_workbook(io.BytesIO(apsx_bytes), data_only=True)
        ws_tmp = wb_tmp.active
        _, _, dstart, dend = analyze_source(ws_tmp)
        mr_tmp, cont_tmp = build_merge_resolver(ws_tmp)
        def gv(r,c): return mr_tmp.get((r,c), ws_tmp.cell(r,c).value)

        receipt_items = []
        for r in range(dstart, dend + 1):
            if r in cont_tmp: continue
            re_num = gv(r, 1); y_val = pf(gv(r, 24)); pay_ch = gv(r, 22)
            if re_num and y_val and y_val > 0:
                receipt_items.append((re_num, y_val, pay_ch))

    bank_match, unmatched = auto_match(receipt_items, all_credits)
    override_parsed = parse_override(override_dict)
    bank_match.update(override_parsed)
    unmatched_final = [u for u in unmatched if (u['re'], u['amount']) not in bank_match]

    with st.spinner("Building Excel output..."):
        out_bytes, matched, blank, total_rows, cont_rows = build_excel(
            apsx_bytes, bank_match, yyyymm)

    out_name = f"{yyyymm}_BKFL_Receipt Report.xlsx"

    st.header("3. ผลลัพธ์")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Data rows", total_rows)
    c2.metric("Sub-item rows zeroed", cont_rows)
    c3.metric("AF/AG filled ✓", matched)
    c4.metric("Blank (unmatched) ⚠", blank)

    if unmatched_final:
        st.warning(f"⚠ **{len(unmatched_final)} รายการยังไม่ match** — AF/AG จะว่าง")
        rows_data = []
        for u in unmatched_final:
            key = f"{u['re']}_{u['amount']}" if u['amount'] == int(u['amount']) else f"{u['re']}_{u['amount']}"
            rows_data.append({
                'RE#': u['re'],
                'ยอดชำระ': u['amount'],
                'ช่องทาง': u['channel'],
                'ประเภทบัญชี': u['bank_type'],
                'ยอดสุทธิ (ประมาณ)': round(u['expected_net'], 2),
                'Key สำหรับ Override': key
            })
        st.dataframe(rows_data, use_container_width=True)
        st.caption(
            "💡 คัดลอก Key จากตารางด้านบน → วางใน Override box → กด Process อีกครั้ง\n\n"
            "Format: `\"KEY\": [ยอดจริงจาก bank statement, \"YYYY-MM-DD\", \"BANK_TYPE\"]`"
        )
    else:
        st.success("✅ ทุกรายการ match ครบ (ยกเว้น cash และ cross-month settlements)")

    st.download_button(
        label=f"⬇ ดาวน์โหลด {out_name}",
        data=out_bytes,
        file_name=out_name,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type='primary',
        use_container_width=True,
    )

# ── FOOTER ────────────────────────────────────────────────────────────────────
st.divider()
st.caption("BKFL Receipt Processor · อัปโหลดไฟล์ APSx + Bank PDFs → Excel คำนวณ DF/Commission")
